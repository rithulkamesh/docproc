"""Excel (.xlsx, .xls) spreadsheet loader."""

from pathlib import Path
from typing import Iterator

from docproc.doc.loaders.base import DocumentLoader, LoadedPage
from docproc.doc.regions import BoundingBox, Region, RegionType


class XLSXLoader(DocumentLoader):
    """Load Excel files via openpyxl."""

    def load(self, path: Path) -> Iterator[LoadedPage]:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet_idx, sheet in enumerate(wb.worksheets):
                rows = []
                regions = []
                for row in sheet.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    row_text = " | ".join(cells).strip()
                    if row_text:
                        rows.append(row_text)
                        regions.append(
                            Region(
                                region_type=RegionType.TABLE,
                                bbox=BoundingBox(x1=0, y1=0, x2=0, y2=0),
                                confidence=0.95,
                                content=row_text,
                                metadata={"sheet": sheet.title, "sheet_idx": sheet_idx},
                            )
                        )
                yield LoadedPage(
                    page_num=sheet_idx,
                    text="\n".join(rows),
                    regions=regions,
                    raw_images=[],
                )
        finally:
            wb.close()

    def get_full_text(self, path: Path) -> str:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            parts = []
            for sheet in wb.worksheets:
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    row_text = " | ".join(cells).strip()
                    if row_text:
                        rows.append(row_text)
                if rows:
                    parts.append(f"=== {sheet.title} ===\n" + "\n".join(rows))
            return "\n\n".join(parts)
        finally:
            wb.close()
